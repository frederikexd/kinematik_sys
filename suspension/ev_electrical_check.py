"""
ev_electrical_check.py — Electrical feasibility gate for KinematiK
====================================================================
The electrics lead identified a real problem: the lap-time simulator can produce
target times that are *physically impossible* for the electrical system to support.
A 47-second autocross might look tight but achievable on paper — but if it demands
sustained current the fuse will blow in, the pack voltage collapses under, or the
pack simply runs dry before crossing the finish line, the lap doesn't happen.

This module bridges the gap.  Given:
  - A simulated lap result (speed profile vs. distance / time)
  - The electrical system parameters loaded from the electrics lead's Excel workbook

…it computes, at each point along the lap:
  1. Required mechanical power at the wheels          (P_wheel, kW)
  2. Required electrical power from the pack          (P_elec = P_wheel / η_drivetrain)
  3. Required pack current                            (I_pack = P_elec / V_pack, A)
  4. Required current per cell in parallel strings    (I_cell = I_pack / n_parallel, A)

Then it checks three hard limits from the Excel spec:
  ├─ Fuse limit:   peak I_pack  ≤  Fuse Max (A)
  ├─ Cell limit:   peak I_cell  ≤  rated cell current (Capacity / nominal discharge C)
  └─ Energy limit: total energy drawn  ≤  usable pack energy (kWh)

And computes the *maximum physically achievable lap time* given the fuse/current
cap — i.e. the lap time that the electrical system can actually sustain, not the
one the mechanical solver found.

Usage
-----
>>> from suspension.ev_electrical_check import ElecParams, check_lap_electrical
>>> params = ElecParams.from_excel("path/to/FSAE_EV_Power_Draw.xlsx")
>>> result = check_lap_electrical(lap_result, params, drivetrain_eff=0.90)
>>> if result.fuse_blown:
...     st.error(result.fuse_message)

All computation is pure Python / NumPy — no Streamlit import here so the module
is usable in unit tests and in the Streamlit UI alike.
"""

from __future__ import annotations

import math
from dataclasses import dataclass, field
from typing import Optional, Sequence

import numpy as np

# ─────────────────────────────────────────────────────────────────────────────
# Data classes
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ElecParams:
    """Electrical system parameters parsed from the electrics lead's Excel sheet."""

    # ── Pack topology ─────────────────────────────────────────────────────────
    pack_voltage_v: float = 504.0          # nominal pack voltage (V)
    pack_capacity_ah: float = 15.0         # pack capacity (Ah)
    n_parallel: int = 3                    # parallel cell strings
    n_series: int = 140                    # series cells per string
    cell_capacity_ah: float = 5.0          # individual cell capacity (Ah)
    cell_r_ohm: float = 0.0128            # cell internal resistance (Ω)

    # ── Limits ────────────────────────────────────────────────────────────────
    fuse_max_a: float = 50.0              # pack-level fuse rating (A)
    nominal_cell_current_a: float = 16.667  # rated current per cell (A) from Excel

    # ── Motor / inverter ─────────────────────────────────────────────────────
    motor_peak_power_kw: float = 150.0    # motor nameplate peak power (kW)
    motor_peak_torque_nm: float = 120.0   # motor nameplate peak torque (N·m)
    motor_efficiency: float = 0.9545      # motor efficiency (fraction)
    motor_max_dc_v: float = 925.0         # motor max DC bus voltage (V)
    wheel_diameter_in: float = 18.0       # wheel diameter (inches)

    # ── Derived / cached ─────────────────────────────────────────────────────
    # computed on post_init
    pack_energy_kwh: float = field(init=False)
    usable_energy_kwh: float = field(init=False)   # 92% usable by default
    max_power_from_fuse_kw: float = field(init=False)

    def __post_init__(self):
        # Pack energy: V × Ah  → Wh → kWh
        self.pack_energy_kwh = (self.pack_voltage_v * self.pack_capacity_ah) / 1000.0
        self.usable_energy_kwh = self.pack_energy_kwh * 0.92   # standard 92% usable
        # The fuse caps instantaneous pack current; convert to a power ceiling
        self.max_power_from_fuse_kw = (self.fuse_max_a * self.pack_voltage_v) / 1000.0

    # ── Factory ──────────────────────────────────────────────────────────────
    @classmethod
    def from_excel(cls, path: str) -> "ElecParams":
        """
        Parse an ElecParams from the electrics lead's Excel workbook.

        Reads two sheets:
          'Battery Pack Calcs'  — pack topology, fuse, cell specs
          'ElecPropulsion'      — motor peak power, torque, efficiency, voltage, wheel size

        Falls back to sensible FSAE-EV defaults on any missing key.
        """
        try:
            import openpyxl  # type: ignore
        except ImportError:
            return cls()  # can't parse, return defaults

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return cls()

        # --- Battery Pack Calcs sheet ---
        pack: dict = {}
        if "Battery Pack Calcs" in wb.sheetnames:
            for row in wb["Battery Pack Calcs"].iter_rows(values_only=True):
                if row[0] and row[1] is not None:
                    pack[str(row[0]).strip()] = row[1]

        # --- ElecPropulsion sheet (first 12 rows of labelled params) ---
        ep: dict = {}
        if "ElecPropulsion" in wb.sheetnames:
            for row in list(wb["ElecPropulsion"].iter_rows(max_row=12, values_only=True)):
                if row[0] and row[1] is not None:
                    ep[str(row[0]).strip()] = row[1]

        def g(d, *keys, default=None):
            """Grab first matching key from dict, ignoring trailing spaces."""
            for k in keys:
                for dk in d:
                    if dk.strip().lower() == k.strip().lower():
                        v = d[dk]
                        try:
                            return float(v)
                        except (TypeError, ValueError):
                            return default
            return default

        n_par  = int(g(pack, "Parrallel Battery Count", "Parallel Battery Count") or 3)
        n_ser  = int(g(pack, "Series Battery Count") or 140)
        cell_c = g(pack, "Capacity Battery Cell (Ah)") or 5.0
        cell_r = g(pack, "Internal Resistance Battery Cell (Ohms)") or 0.0128
        v_pack = g(pack, "Battery Pack Nominal Voltage (V)") or (n_ser * 3.6)
        cap_ah = g(pack, "Pack Capacity (Ah)") or (cell_c * n_par)
        fuse   = g(pack, "Fuse Max (A)") or 50.0
        i_cell = g(pack, "Current Across a Single Cell (A)") or (cap_ah / n_par)

        mot_pk_kw  = g(ep, "Motor Peak Power (kW)") or 150.0
        mot_pk_nm  = g(ep, "Motor Peak Torque (Nm)") or 120.0
        mot_eff    = g(ep, "Motor Efficiency", "Motor Efficiency ") or 0.9545
        mot_v      = g(ep, "Motor Max DC Voltage (V)") or 925.0
        whl_in     = g(ep, "Wheel Diameter (in)") or 18.0

        return cls(
            pack_voltage_v=v_pack,
            pack_capacity_ah=cap_ah,
            n_parallel=n_par,
            n_series=n_ser,
            cell_capacity_ah=cell_c,
            cell_r_ohm=cell_r,
            fuse_max_a=fuse,
            nominal_cell_current_a=i_cell,
            motor_peak_power_kw=mot_pk_kw,
            motor_peak_torque_nm=mot_pk_nm,
            motor_efficiency=mot_eff,
            motor_max_dc_v=mot_v,
            wheel_diameter_in=whl_in,
        )

    @classmethod
    def from_excel_bytes(cls, file_bytes: bytes) -> "ElecParams":
        """Parse from an in-memory file (e.g. st.file_uploader)."""
        import io
        try:
            import openpyxl
        except ImportError:
            return cls()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception:
            return cls()
        # write to a temp file and re-use from_excel (simplest)
        import tempfile, os
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(file_bytes)
        tmp.close()
        try:
            result = cls.from_excel(tmp.name)
        finally:
            os.unlink(tmp.name)
        return result


@dataclass
class ElecCheckResult:
    """
    Full electrical feasibility report for one simulated lap.

    Flags
    -----
    fuse_blown       True if peak pack current exceeds the fuse rating
    cell_overcurrent True if any individual cell exceeds its rated current
    energy_empty     True if the lap drains the pack below the usable floor
    ok               True only when *all three* checks pass

    Numbers
    -------
    peak_current_a       Maximum pack current seen anywhere on lap (A)
    peak_current_speed   Speed at which that peak occurred (m/s)
    avg_current_a        Mean pack current over the traction phase (A)
    energy_used_kwh      Total electrical energy drawn from the pack this lap (kWh)
    usable_energy_kwh    Pack's usable energy ceiling for comparison (kWh)
    max_safe_speed_ms    Highest speed the electrical system can sustain continuously
    max_safe_power_kw    Fuse-limited power ceiling (kW)
    current_profile_a    np.ndarray — pack current at each lap sample (A)
    power_profile_kw     np.ndarray — electrical power at each lap sample (kW)
    speed_profile_ms     np.ndarray — speed at each sample (m/s)

    Messages
    --------
    fuse_message         Human-readable fuse verdict
    cell_message         Human-readable per-cell verdict
    energy_message       Human-readable energy verdict
    summary              One-line overall verdict
    warnings             List of additional engineering notes
    """

    # ── Flags ─────────────────────────────────────────────────────────────────
    ok: bool = True
    fuse_blown: bool = False
    cell_overcurrent: bool = False
    energy_empty: bool = False

    # ── Current / power numbers ────────────────────────────────────────────────
    peak_current_a: float = 0.0
    peak_current_speed_ms: float = 0.0
    avg_current_a: float = 0.0
    energy_used_kwh: float = 0.0
    usable_energy_kwh: float = 0.0
    max_safe_speed_ms: float = 0.0
    max_safe_power_kw: float = 0.0
    fuse_max_a: float = 50.0

    # ── Per-cell numbers ──────────────────────────────────────────────────────
    peak_cell_current_a: float = 0.0
    cell_current_limit_a: float = 0.0

    # ── Profiles (arrays) ─────────────────────────────────────────────────────
    current_profile_a: np.ndarray = field(default_factory=lambda: np.array([]))
    power_profile_kw: np.ndarray  = field(default_factory=lambda: np.array([]))
    speed_profile_ms: np.ndarray  = field(default_factory=lambda: np.array([]))

    # ── Messages ──────────────────────────────────────────────────────────────
    fuse_message: str = ""
    cell_message: str = ""
    energy_message: str = ""
    summary: str = ""
    warnings: list = field(default_factory=list)


# ─────────────────────────────────────────────────────────────────────────────
# Core computation
# ─────────────────────────────────────────────────────────────────────────────

def check_lap_electrical(
    speed_ms: Sequence[float],
    distance_m: Sequence[float],
    params: ElecParams,
    *,
    drivetrain_eff: float = 0.90,
    vehicle_mass_kg: float = 230.0,
    drag_cda: float = 1.10,
    crr: float = 0.018,
    rho_air: float = 1.225,
) -> ElecCheckResult:
    """
    Check whether the simulated lap speed profile is electrically achievable.

    Parameters
    ----------
    speed_ms      : Speed at each sample point (m/s).  Can be a lap result's `.v`.
    distance_m    : Distance at each sample point (m). Can be a lap result's `.s`.
    params        : ElecParams parsed from the electrics lead's Excel workbook.
    drivetrain_eff: Combined inverter + motor + gearbox efficiency (fraction).
    vehicle_mass_kg: Total vehicle mass including driver (kg).
    drag_cda      : Aerodynamic drag area (m²).  Defaults to FSAE-typical value.
    crr           : Rolling resistance coefficient.
    rho_air       : Air density (kg/m³).

    Returns
    -------
    ElecCheckResult with all flags, numbers, profiles, and human-readable messages.
    """
    v = np.asarray(speed_ms, dtype=float)
    s = np.asarray(distance_m, dtype=float)

    if len(v) < 2:
        r = ElecCheckResult()
        r.ok = False
        r.summary = "Not enough speed data to run the electrical check."
        return r

    # ── Step 1: compute required tractive force at each point ─────────────────
    # F_drag = ½ ρ CdA v²
    F_drag = 0.5 * rho_air * drag_cda * v ** 2
    # F_roll = crr × mg
    F_roll = crr * vehicle_mass_kg * 9.81
    # F_accel = m × a  (finite difference; clamp to 0 — we only care about traction)
    dv = np.gradient(v, s)          # dv/ds (m/s per m)
    a  = v * dv                     # a = v × dv/ds  (m/s²)
    F_accel = vehicle_mass_kg * np.maximum(a, 0.0)  # only acceleration phases

    F_total = F_drag + F_roll + F_accel    # N, at each point

    # ── Step 2: mechanical power at wheels ────────────────────────────────────
    P_wheel_w = F_total * np.maximum(v, 0.5)          # W
    P_wheel_kw = P_wheel_w / 1000.0

    # ── Step 3: electrical power from pack ────────────────────────────────────
    eff = max(drivetrain_eff, 0.50)
    P_elec_kw = P_wheel_kw / eff                      # kW drawn from pack
    P_elec_w  = P_elec_kw * 1000.0                    # W

    # ── Step 4: pack current ──────────────────────────────────────────────────
    V_pack = params.pack_voltage_v
    I_pack = P_elec_w / max(V_pack, 1.0)              # A drawn from pack

    # ── Step 5: per-cell current ──────────────────────────────────────────────
    I_cell = I_pack / max(params.n_parallel, 1)       # A per cell

    # ── Step 6: energy integral (trapezoidal rule) ────────────────────────────
    # ΔE = P × Δt;  Δt = Δs / v_avg
    dt = np.zeros(len(v))
    for i in range(1, len(v)):
        v_avg = (v[i - 1] + v[i]) / 2.0
        ds = abs(s[i] - s[i - 1])
        dt[i] = ds / max(v_avg, 0.5)

    energy_per_step_wh = P_elec_w * dt / 3600.0      # Wh at each step
    total_energy_kwh = float(np.sum(energy_per_step_wh)) / 1000.0

    # ── Step 7: limits ────────────────────────────────────────────────────────
    peak_I     = float(np.max(I_pack))
    peak_speed = float(v[int(np.argmax(I_pack))])
    _traction_mask = I_pack > 0.1
    avg_I      = float(np.mean(I_pack[_traction_mask])) if _traction_mask.any() else 0.0

    peak_I_cell  = float(np.max(I_cell))
    fuse_limit   = params.fuse_max_a
    cell_limit   = params.nominal_cell_current_a

    fuse_blown       = peak_I > fuse_limit
    cell_overcurrent = peak_I_cell > cell_limit
    energy_empty     = total_energy_kwh > params.usable_energy_kwh

    # ── Step 8: maximum safe continuous speed (fuse-limited) ──────────────────
    # At constant speed: P_wheel = F_drag × v = ½ρCdA v³ + crr·mg·v
    # P_elec = P_wheel / eff  ≤  fuse_limit × V_pack
    max_safe_P_kw = params.max_power_from_fuse_kw
    max_safe_P_w  = max_safe_P_kw * 1000.0

    # Solve  ½ρCdA v³ + crr·mg·v  =  max_safe_P_w × eff   for v  (numerically)
    v_test = np.linspace(0.1, 100.0, 5000)
    P_test = (0.5 * rho_air * drag_cda * v_test ** 3
              + crr * vehicle_mass_kg * 9.81 * v_test)
    P_cap  = max_safe_P_w * eff
    idx    = np.searchsorted(P_test, P_cap)
    max_safe_speed_ms = float(v_test[min(idx, len(v_test) - 1)])

    # ── Step 9: assemble result ───────────────────────────────────────────────
    ok = not (fuse_blown or cell_overcurrent or energy_empty)

    # Fuse message
    if fuse_blown:
        margin = peak_I - fuse_limit
        fuse_msg = (
            f"🚨 FUSE WILL BLOW — peak pack current {peak_I:.1f} A exceeds "
            f"the {fuse_limit:.0f} A fuse by {margin:.1f} A "
            f"(at {peak_speed * 3.6:.1f} km/h). "
            f"The lap-time sim assumed {max_safe_P_kw:.1f} kW available; "
            f"the fuse actually caps you at {max_safe_P_kw:.1f} kW "
            f"(~{max_safe_speed_ms * 3.6:.1f} km/h sustained). "
            f"Slow your roll — or upsize the fuse with the electrics lead."
        )
    else:
        headroom = fuse_limit - peak_I
        fuse_msg = (
            f"✅ Fuse OK — peak {peak_I:.1f} A vs {fuse_limit:.0f} A limit "
            f"({headroom:.1f} A headroom)."
        )

    # Cell current message
    if cell_overcurrent:
        cell_msg = (
            f"⚠️ Cell overcurrent — each cell sees {peak_I_cell:.1f} A peak "
            f"(limit {cell_limit:.1f} A across {params.n_parallel} parallel strings). "
            f"This degrades pack life and risks thermal runaway."
        )
    else:
        cell_msg = (
            f"✅ Cell current OK — {peak_I_cell:.1f} A peak per cell "
            f"(rated {cell_limit:.1f} A)."
        )

    # Energy message
    if energy_empty:
        deficit = total_energy_kwh - params.usable_energy_kwh
        energy_msg = (
            f"🚨 PACK RUNS DRY — this lap draws {total_energy_kwh:.3f} kWh, "
            f"but only {params.usable_energy_kwh:.3f} kWh is usable "
            f"({deficit:.3f} kWh short). The car won't finish."
        )
    else:
        pct_used = (total_energy_kwh / params.usable_energy_kwh * 100.0
                    if params.usable_energy_kwh > 0 else 0.0)
        energy_msg = (
            f"✅ Energy OK — lap uses {total_energy_kwh:.3f} kWh "
            f"({pct_used:.1f}% of {params.usable_energy_kwh:.3f} kWh usable pack)."
        )

    # Summary
    issues = []
    if fuse_blown:
        issues.append("fuse over-current")
    if cell_overcurrent:
        issues.append("cell over-current")
    if energy_empty:
        issues.append("energy deficit")
    if ok:
        summary = (
            f"✅ Electrically feasible — lap is within fuse, cell, and energy limits. "
            f"Peak draw {peak_I:.1f} A / {max_safe_P_kw:.1f} kW fuse ceiling."
        )
    else:
        summary = (
            f"❌ Lap NOT electrically feasible — {', '.join(issues)}. "
            f"Fuse ceiling: {fuse_limit:.0f} A ({max_safe_P_kw:.1f} kW at {V_pack:.0f} V). "
            f"Estimated max sustainable speed: {max_safe_speed_ms * 3.6:.1f} km/h."
        )

    warnings_list = []
    if max_safe_speed_ms < float(np.max(v)):
        max_v_kmh = float(np.max(v)) * 3.6
        warnings_list.append(
            f"Top speed in the sim ({max_v_kmh:.1f} km/h) exceeds the fuse-limited "
            f"sustained ceiling ({max_safe_speed_ms * 3.6:.1f} km/h). "
            f"Brief bursts may be OK (capacitive support), but sustained speed is capped."
        )
    if params.pack_capacity_ah < 10.0:
        warnings_list.append(
            "Pack capacity is below 10 Ah — double-check the Excel sheet values."
        )

    return ElecCheckResult(
        ok=ok,
        fuse_blown=fuse_blown,
        cell_overcurrent=cell_overcurrent,
        energy_empty=energy_empty,
        peak_current_a=peak_I,
        peak_current_speed_ms=peak_speed,
        avg_current_a=avg_I,
        energy_used_kwh=total_energy_kwh,
        usable_energy_kwh=params.usable_energy_kwh,
        max_safe_speed_ms=max_safe_speed_ms,
        max_safe_power_kw=max_safe_P_kw,
        fuse_max_a=fuse_limit,
        peak_cell_current_a=peak_I_cell,
        cell_current_limit_a=cell_limit,
        current_profile_a=I_pack,
        power_profile_kw=P_elec_kw,
        speed_profile_ms=v,
        fuse_message=fuse_msg,
        cell_message=cell_msg,
        energy_message=energy_msg,
        summary=summary,
        warnings=warnings_list,
    )


def check_lap_from_speed_csv(
    speed_mph: Sequence[float],
    time_s: Sequence[float],
    params: ElecParams,
    *,
    drivetrain_eff: float = 0.90,
    vehicle_mass_kg: float = 230.0,
    drag_cda: float = 1.10,
    crr: float = 0.018,
) -> ElecCheckResult:
    """
    Run the electrical check from the SpeedVsTime CSV data in the Excel workbook
    (columns: time_s, speed_mph).

    Converts speed to m/s, integrates to get distance, then delegates to
    `check_lap_electrical`.
    """
    t = np.asarray(time_s, dtype=float)
    v_mph = np.asarray(speed_mph, dtype=float)
    v_ms = v_mph * 0.44704                          # mph → m/s

    # Integrate speed over time to get distance
    s = np.zeros(len(t))
    for i in range(1, len(t)):
        dt = t[i] - t[i - 1]
        s[i] = s[i - 1] + 0.5 * (v_ms[i - 1] + v_ms[i]) * dt

    return check_lap_electrical(
        v_ms, s, params,
        drivetrain_eff=drivetrain_eff,
        vehicle_mass_kg=vehicle_mass_kg,
        drag_cda=drag_cda,
        crr=crr,
    )


def load_speed_vs_time_from_bytes(file_bytes: bytes) -> tuple[np.ndarray, np.ndarray]:
    """
    Load the SpeedVsTime sheet from in-memory Excel bytes.

    Returns (time_s, speed_mph) arrays.

    Raises
    ------
    ValueError / ImportError  propagated from load_speed_vs_time on bad data.
    """
    import tempfile, os
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(file_bytes)
    tmp.close()
    try:
        result = load_speed_vs_time(tmp.name)
    finally:
        os.unlink(tmp.name)
    return result


def load_speed_vs_time(path: str) -> tuple[np.ndarray, np.ndarray]:
    """
    Load the SpeedVsTime sheet from the electrics lead's Excel workbook.

    Returns
    -------
    (time_s, speed_mph)  — both as float arrays, header row stripped.

    Raises
    ------
    ValueError  if the sheet is missing or contains fewer than 2 numeric rows.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required to read Excel files. Install it with: pip install openpyxl")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not open Excel workbook: {exc}") from exc

    if "SpeedVsTime" not in wb.sheetnames:
        available = ", ".join(wb.sheetnames) or "(none)"
        raise ValueError(
            f"No 'SpeedVsTime' sheet found in the workbook. "
            f"Available sheets: {available}. "
            f"Please add a sheet named exactly 'SpeedVsTime' with columns: time_s, speed_mph."
        )

    times, speeds = [], []
    for row in wb["SpeedVsTime"].iter_rows(values_only=True):
        if len(row) < 2:
            continue
        t, v = row[0], row[1]
        if isinstance(t, (int, float)) and isinstance(v, (int, float)):
            times.append(float(t))
            speeds.append(float(v))

    if len(times) < 2:
        raise ValueError(
            f"SpeedVsTime sheet has only {len(times)} numeric data row(s) — need at least 2. "
            f"Make sure column A = time (s) and column B = speed (mph), with a header row on row 1."
        )

    return np.array(times), np.array(speeds)
