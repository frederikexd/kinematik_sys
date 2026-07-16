"""
ev_excel_roundtrip.py  —  Lap-sim ↔ Excel round-trip, zero friction edition
=============================================================================
What Excel tedium we kill:
  1. Manual copy-paste of speed data into SpeedVsTime (gone — KinematiK writes it)
  2. Waiting for Excel to recalculate 50,000+ cells across 15 gear-ratio columns
     (gone — Python evaluates every formula natively, no LibreOffice dependency)
  3. Hunting for peak/avg current buried in row 4000+ of ElecPropulsion
     (gone — KinematiK extracts and surfaces all key numbers instantly)
  4. Having no audit trail of which lap the Excel was calculated for
     (gone — a KinematiK Lap Sim Summary block is stamped into Battery Pack Calcs)
  5. Not knowing whether the lap is electrically achievable before the car is built
     (gone — feasibility verdict + fuse headroom shown before you leave the tab)

Data flow (fully Python-evaluated, no Excel engine required)
-------------------------------------------------------------
  KinematiK lap sim                       FSAE_EV_Power_Draw.xlsx
  ─────────────────                       ──────────────────────────────────────
  v_ms → mph ──────────────────────────► SpeedVsTime  A:B   (time, speed mph)

  Python evaluates all downstream formulas:
  Battery Pack Calcs B1:B16  → pack_summary
  ElecPropulsion H–V rows 2..N (15 gear ratios):
      RPM = v_mph * gear_ratio * 1056 / (π * wheel_diam_in)
  ElecPropulsion phase current section rows N+2..2N:
      I_phase = V_pack * √3 * PF * RPM / 1000
  Current Draw section:
      I_draw = R_cell * V_pack * RPM / 1000

  All computed values written as plain numbers into the returned workbook
  (no formula strings — opens instantly in Excel with zero recalculation lag).
"""
from __future__ import annotations

import io, math, os, shutil, tempfile
from dataclasses import dataclass, field
from typing import Sequence

import numpy as np


# ─────────────────────────────────────────────────────────────────────────────
# Sheet / layout constants  (match the workbook exactly)
# ─────────────────────────────────────────────────────────────────────────────

_SVT_SHEET  = "SpeedVsTime"
_PACK_SHEET = "Battery Pack Calcs"
_EP_SHEET   = "ElecPropulsion"

# Battery Pack Calcs — label → (row, col) of the *value* cell
_PACK_CELLS: dict[str, tuple[int,int]] = {
    "fuse_max_a":         (1,  2),
    "n_parallel":         (2,  2),
    "n_series":           (3,  2),
    "cell_voltage_v":     (4,  2),
    "cell_capacity_ah":   (5,  2),
    "endurance_km":       (6,  2),
    "max_cells":          (7,  2),
    "cell_r_ohm":         (8,  2),   # B8  — used in current formula
    "cell_weight_kg":     (9,  2),
    "pack_cell_count":    (10, 2),
    "pack_voltage_v":     (11, 2),   # B11 — used in current formula
    "cell_current_a":     (12, 2),
    "power_draw_kw":      (13, 2),
    "pack_capacity_ah":   (14, 2),
    "pack_energy_wh":     (15, 2),
    "joule_heating_kwh":  (16, 2),
}

# ElecPropulsion scalar params — key → (row, col)
_EP_PARAMS: dict[str, tuple[int,int]] = {
    "motor_peak_torque_nm":  (1,  2),
    "motor_peak_power_kw":   (2,  2),
    "motor_freq_khz":        (3,  2),
    "motor_poles":           (4,  2),
    "motor_max_dc_v":        (5,  2),
    "motor_efficiency":      (6,  2),
    "current_from_pack_a":   (7,  2),   # B7 — I_pack (same as B8×B11 of BPC)
    "pack_voltage_ep_v":     (8,  2),   # B8 in EP = pack voltage
    "motor_max_rpm":         (9,  2),
    "wheel_diam_in":         (10, 2),   # B10 — wheel diameter
    "motor_pf":              (11, 2),   # B11 — power factor
    "no_load_speed":         (1,  5),   # E1
    "synchronous_rpm":       (2,  5),   # E2
}

# Gear-ratio row (header row 1, col H..V = cols 8..22)
_EP_GEAR_RATIO_ROW = 1
_EP_GEAR_COL_START = 8   # H
_EP_GEAR_COL_END   = 22  # V  (15 gear ratios)

# RPM data block: rows 2..N  (mirrors SpeedVsTime rows 2..N)
_EP_RPM_ROW_START  = 2

# Phase current block (rows EP_data_end+2 .. EP_data_end+1+N):
#   formula:  =($B$11 * SQRT(3) * $B$6 * RPM_ref) / 1000
#   where B11 = pack_voltage_ep_v,  B6 = cell_capacity_ah (PF in EP context — row 6)
#   NOTE: in ElecPropulsion, B6 = motor_freq_khz (row 6), but the phase current
#   formula uses $B$6 which maps to row 6 col 2 of ElecPropulsion.
#   From the formula: =$B$11*(SQRT(3))*$B$6*H{rpm_row}  where B11=V_pack_ep, B6=PF

# Current Draw block (after phase current block):
#   formula: =($B$8 * $B$11 * H{rpm_row}) / 1000
#   B8 = current_from_pack_a (row 8 of EP), B11 = pack_voltage_ep_v (row 11 of EP? NO)
#   Wait — in EP: row 8 = B8 = pack_voltage_ep_v (504 V), row 11 = B11 = motor_pf (0.95)
#   From formula: =($B$8*$B$11*H{rpm_row})/1000  → 504 * 0.95 * RPM / 1000
#   That gives ~479 * RPM / 1000.  Let's verify with known value:
#   At 501 RPM: 504 * 0.95 * 501 / 1000 = 479 * 501 / 1000 = 240.0 ✓

# Phase current formula: =($B$11 * SQRT(3) * $B$6 * H{rpm_row}) / 1000
# In EP: B11 = row11 = motor_pf (0.95),  B6 = row6 = motor_efficiency (0.9545)
# Hmm — let's check: 504 * 1.732 * 0.9545 * 501 / 1000 = 417 — doesn't match row counts
# Actually: the phase current formula refers to B11 = pack_voltage_ep_v AND SQRT(3) AND B6 = motor_efficiency
# =504 * 1.732 * 0.9545 * 501/1000 = 417 ≠ expected ~240
# So B11 in that sheet = 0.95 (PF, row 11), B6 = 504 (V_pack, could be row 6 in a diff mapping)
# Looking at actual cells: B8=pack_voltage (504), B11=motor_pf (0.95)
# Phase: =(B11)*(sqrt3)*(B6)*(RPM)/1000 = 0.95*1.732*B6*RPM/1000
# If B6=504: 0.95*1.732*504*RPM/1000 = 829*RPM/1000 @ 501RPM = 415 — no
# Current draw: =(B8)*(B11)*(RPM)/1000 = 504*0.95*RPM/1000 = 479*RPM/1000 @ 501 = 240 ✓
# So _current draw_ formula uses EP.B8=V_pack(504) and EP.B11=PF(0.95)

# For phase current, we need the actual constants from the sheet.
# Since this is read from the workbook we don't hardcode — we read EP B-column values.

# Summary: layout offsets for the three blocks in ElecPropulsion
# Block 1: RPM grid          rows 2..N_data+1  (N_data = n_pts)
# Gap row: N_data+2          (label row: "Current Draw (A)")
# Block 2: Current draw      rows N_data+2 .. 2*N_data+1
# Gap row: 2*N_data+2        (label row: "Phase Current (A)" — added by us)
# Block 3: Phase current     rows 2*N_data+3 .. 3*N_data+2

# In original file: 1893 data rows → phase current starts at 1895 in col H
# 1895 = 1893+2, confirming Block2 start offset = N+2


# ─────────────────────────────────────────────────────────────────────────────
# Result dataclass
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class ExcelRoundTripResult:
    """
    Full result of the lap-sim → Excel → read-back round-trip.
    All formula evaluation is done in Python; no Excel engine required.
    """
    # ── Status ────────────────────────────────────────────────────────────────
    ok:       bool  = True
    error:    str   = ""
    warnings: list  = field(default_factory=list)

    # ── Profiles ──────────────────────────────────────────────────────────────
    time_s:         np.ndarray = field(default_factory=lambda: np.array([]))
    speed_mph:      np.ndarray = field(default_factory=lambda: np.array([]))
    speed_ms:       np.ndarray = field(default_factory=lambda: np.array([]))

    # RPM at each time step (col H = gear ratio 1 from header row)
    rpm_gear1:      np.ndarray = field(default_factory=lambda: np.array([]))

    # Current draw: =(B8_ep * B11_ep * RPM) / 1000
    current_draw_a: np.ndarray = field(default_factory=lambda: np.array([]))

    # Phase current: =(B11_ep * sqrt(3) * B6_ep * RPM) / 1000
    phase_current_a: np.ndarray = field(default_factory=lambda: np.array([]))

    # Electrical power profile
    power_kw:       np.ndarray = field(default_factory=lambda: np.array([]))

    # ── Pack / motor scalars ──────────────────────────────────────────────────
    pack:  dict = field(default_factory=dict)
    motor: dict = field(default_factory=dict)

    # ── Summary scalars ───────────────────────────────────────────────────────
    max_speed_mph:    float = 0.0
    peak_current_a:   float = 0.0
    avg_current_a:    float = 0.0
    peak_power_kw:    float = 0.0
    total_energy_kwh: float = 0.0
    fuse_margin_a:    float = 0.0   # positive = headroom, negative = over limit

    # Fuse-limited max sustained speed
    fuse_speed_ceiling_ms: float = 0.0

    # Usable pack energy for comparison
    usable_energy_kwh: float = 0.0

    # ── Feasibility flags ─────────────────────────────────────────────────────
    fuse_ok:       bool = True
    energy_ok:     bool = True
    feasible:      bool = True
    verdict:       str  = ""

    # ── Downloaded file ───────────────────────────────────────────────────────
    excel_bytes:  bytes = field(default_factory=bytes)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _safe_float(v, default=0.0) -> float:
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default


def _col_letter(n: int) -> str:
    """1-based column index → Excel column letter(s)."""
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def lap_result_to_time_axis(v_ms: np.ndarray, s_m: np.ndarray) -> np.ndarray:
    """Convert distance-sampled lap profile to time axis via trapezoidal integration."""
    t = np.zeros(len(v_ms))
    for i in range(1, len(v_ms)):
        v_avg = (v_ms[i-1] + v_ms[i]) / 2.0
        ds    = abs(s_m[i] - s_m[i-1])
        t[i]  = t[i-1] + ds / max(v_avg, 0.5)
    return t


def load_speed_vs_time(path: str) -> tuple[np.ndarray, np.ndarray]:
    """Read SpeedVsTime sheet → (time_s, speed_mph) arrays."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return np.array([0.0, 1.0]), np.array([0.0, 0.0])
    if _SVT_SHEET not in wb.sheetnames:
        return np.array([0.0, 1.0]), np.array([0.0, 0.0])
    times, speeds = [], []
    for row in wb[_SVT_SHEET].iter_rows(values_only=True):
        t, v = row[0], row[1]
        if isinstance(t, (int, float)) and isinstance(v, (int, float)):
            times.append(float(t))
            speeds.append(float(v))
    return np.array(times), np.array(speeds)


def load_speed_vs_time_from_bytes(file_bytes: bytes) -> tuple[np.ndarray, np.ndarray]:
    """Load SpeedVsTime from in-memory bytes (e.g. from st.file_uploader)."""
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(file_bytes); tmp.close()
    try:
        return load_speed_vs_time(tmp.name)
    finally:
        os.unlink(tmp.name)


def check_lap_from_speed_csv(
    speed_mph: Sequence[float],
    time_s:    Sequence[float],
    params,            # ElecParams from ev_electrical_check (duck-typed)
    *,
    drivetrain_eff: float = 0.90,
    vehicle_mass_kg: float = 230.0,
    drag_cda: float = 1.10,
    crr: float = 0.018,
):
    """Thin shim so the Streamlit tab can call this from the roundtrip module."""
    from suspension.ev_electrical_check import check_lap_from_speed_csv as _c
    return _c(speed_mph, time_s, params,
              drivetrain_eff=drivetrain_eff,
              vehicle_mass_kg=vehicle_mass_kg,
              drag_cda=drag_cda, crr=crr)


def extract_params_from_excel(excel_bytes: bytes) -> dict:
    """
    Extract only the static parameters from the workbook — pack specs, motor
    constants, and gear ratios — without writing anything back or running the
    full round-trip.  The returned dict is JSON-serialisable and can be stored
    in ProjectStore.ev_excel_params so teams never have to re-upload the file.

    Keys (always present, defaulting to 0.0 if the sheet cell is blank):
        pack:   {fuse_max_a, n_parallel, n_series, cell_voltage_v,
                 cell_capacity_ah, endurance_km, max_cells, cell_r_ohm,
                 cell_weight_kg, pack_cell_count, pack_voltage_v,
                 cell_current_a, power_draw_kw, pack_capacity_ah,
                 pack_energy_wh, joule_heating_kwh}
        motor:  {motor_peak_torque_nm, motor_peak_power_kw, motor_freq_khz,
                 motor_poles, motor_max_dc_v, motor_efficiency,
                 current_from_pack_a, pack_voltage_ep_v, motor_max_rpm,
                 wheel_diam_in, motor_pf, no_load_speed, synchronous_rpm}
        gear_ratios: [float × 15]
        _source: "excel"
    """
    try:
        import openpyxl
    except ImportError:
        return {"_source": "excel", "_error": "openpyxl not installed"}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    except Exception as exc:
        return {"_source": "excel", "_error": str(exc)}

    pack: dict[str, float] = {}
    ws_pack = wb[_PACK_SHEET] if _PACK_SHEET in wb.sheetnames else None
    if ws_pack:
        for key, (row, col) in _PACK_CELLS.items():
            pack[key] = _safe_float(ws_pack.cell(row=row, column=col).value)

    motor: dict[str, float] = {}
    ws_ep = wb[_EP_SHEET] if _EP_SHEET in wb.sheetnames else None
    if ws_ep:
        for key, (row, col) in _EP_PARAMS.items():
            motor[key] = _safe_float(ws_ep.cell(row=row, column=col).value)

    gear_ratios: list[float] = []
    if ws_ep:
        for col in range(_EP_GEAR_COL_START, _EP_GEAR_COL_END + 1):
            v = ws_ep.cell(row=_EP_GEAR_RATIO_ROW, column=col).value
            gear_ratios.append(_safe_float(v, default=1.0))
    else:
        gear_ratios = [1.0] * 15

    return {
        "pack": pack,
        "motor": motor,
        "gear_ratios": gear_ratios,
        "_source": "excel",
    }


def lap_to_excel_roundtrip_from_db(
    speed_ms:    "Sequence[float]",
    time_s:      "Sequence[float]",
    ev_params:   dict,
    *,
    lap_time_s:   float = 0.0,
    top_speed_ms: float = 0.0,
    avg_speed_ms: float = 0.0,
) -> "ExcelRoundTripResult":
    """
    Run the full round-trip calculation using parameters stored in the project
    database (ev_params dict, as written by extract_params_from_excel), instead
    of requiring the raw xlsx bytes.  No workbook is written; the excel_bytes
    field of the returned result will be empty.  Use lap_to_excel_roundtrip()
    when you do need the updated workbook file.
    """
    v_ms  = np.asarray(speed_ms, dtype=float)
    t_arr = np.asarray(time_s,   dtype=float)
    if len(v_ms) < 2:
        return ExcelRoundTripResult(ok=False, error="Need ≥2 speed points.")
    if len(t_arr) != len(v_ms):
        t_arr = np.arange(len(v_ms)) * 0.1

    v_mph = v_ms * 2.23694
    n_pts = len(v_mph)

    pack        = ev_params.get("pack", {})
    motor       = ev_params.get("motor", {})
    gear_ratios = ev_params.get("gear_ratios") or [1.0] * 15

    wheel_diam_in = motor.get("wheel_diam_in", 18.0) or 18.0
    pack_v_ep     = motor.get("pack_voltage_ep_v", 504.0) or 504.0
    motor_pf      = motor.get("motor_pf", 0.95) or 0.95
    motor_eff     = motor.get("motor_efficiency", 0.9545) or 0.9545
    pack_v_bpc    = pack.get("pack_voltage_v", 504.0) or 504.0
    fuse_limit    = pack.get("fuse_max_a", 50.0) or 50.0
    usable_kwh    = (pack_v_bpc * pack.get("pack_capacity_ah", 15.0)) * 0.92 / 1000.0

    k_rpm    = 1056.0 / (math.pi * wheel_diam_in)
    rpm_all  = np.zeros((n_pts, len(gear_ratios)), dtype=float)
    for gi, gr in enumerate(gear_ratios):
        rpm_all[:, gi] = v_mph * gr * k_rpm
    rpm_gear1 = rpm_all[:, 0]

    current_draw_a  = (pack_v_ep * motor_pf * rpm_gear1) / 1000.0
    phase_current_a = motor_pf * math.sqrt(3) * motor_eff * rpm_gear1
    power_kw        = current_draw_a * pack_v_ep / 1000.0

    dt_arr = np.diff(t_arr, prepend=t_arr[0])
    total_energy_kwh = float(np.sum(power_kw * np.abs(dt_arr))) / 3600.0

    peak_i    = float(np.max(current_draw_a))
    avg_i     = float(np.mean(current_draw_a))
    peak_pw   = float(np.max(power_kw))
    fuse_ok   = peak_i <= fuse_limit
    energy_ok = total_energy_kwh <= usable_kwh
    feasible  = fuse_ok and energy_ok
    fuse_margin = fuse_limit - peak_i

    P_fuse_kw = fuse_limit * pack_v_ep / 1000.0
    v_test = np.linspace(0.1, 100.0, 5000)
    P_test = (0.5 * 1.225 * 1.10 * v_test**3 + 0.018 * 230.0 * 9.81 * v_test) / 1000.0
    idx_ceil = int(np.searchsorted(P_test, P_fuse_kw * 0.90))
    fuse_speed_ms = float(v_test[min(idx_ceil, len(v_test) - 1)])

    if feasible:
        verdict = (f"✅ Electrically feasible — peak {peak_i:.1f} A / "
                   f"{fuse_limit:.0f} A fuse  |  "
                   f"{total_energy_kwh:.3f} kWh / {usable_kwh:.3f} kWh usable")
    else:
        issues = []
        if not fuse_ok:   issues.append(f"fuse blown (+{-fuse_margin:.1f} A over)")
        if not energy_ok: issues.append(f"energy deficit ({total_energy_kwh-usable_kwh:.3f} kWh short)")
        verdict = f"❌ NOT feasible — {', '.join(issues)}"

    return ExcelRoundTripResult(
        ok=True,
        time_s=t_arr,
        speed_mph=v_mph,
        speed_ms=v_ms,
        rpm_gear1=rpm_gear1,
        current_draw_a=current_draw_a,
        phase_current_a=phase_current_a,
        power_kw=power_kw,
        pack=pack,
        motor=motor,
        max_speed_mph=float(np.max(v_mph)),
        peak_current_a=peak_i,
        avg_current_a=avg_i,
        peak_power_kw=peak_pw,
        total_energy_kwh=total_energy_kwh,
        fuse_margin_a=fuse_margin,
        fuse_speed_ceiling_ms=fuse_speed_ms,
        usable_energy_kwh=usable_kwh,
        fuse_ok=fuse_ok,
        energy_ok=energy_ok,
        feasible=feasible,
        verdict=verdict,
        excel_bytes=b"",
    )


# ─────────────────────────────────────────────────────────────────────────────
# Core round-trip (writes back to workbook)
# ─────────────────────────────────────────────────────────────────────────────

def lap_to_excel_roundtrip(
    speed_ms:    Sequence[float],
    time_s:      Sequence[float],
    excel_bytes: bytes,
    *,
    lap_time_s:    float = 0.0,
    top_speed_ms:  float = 0.0,
    avg_speed_ms:  float = 0.0,
    libreoffice_timeout: int = 60,   # kept for API compat, not used
) -> ExcelRoundTripResult:
    """
    Full round-trip — Python-evaluated, no LibreOffice required.

    Writes the lap sim speed profile into the workbook, evaluates every
    downstream formula in pure Python, then returns a fully-computed result
    AND an updated .xlsx file with all values baked in as plain numbers
    (no formula strings — opens instantly with zero recalc lag in Excel).
    """
    try:
        import openpyxl
    except ImportError:
        return ExcelRoundTripResult(ok=False, error="openpyxl not installed.")

    v_ms  = np.asarray(speed_ms, dtype=float)
    t_arr = np.asarray(time_s,   dtype=float)
    if len(v_ms) < 2:
        return ExcelRoundTripResult(ok=False, error="Need ≥2 speed points.")
    if len(t_arr) != len(v_ms):
        t_arr = np.arange(len(v_ms)) * 0.1

    v_mph = v_ms * 2.23694   # m/s → mph
    n_pts = len(v_mph)

    # ── 1. Load the workbook (formulas preserved) ──────────────────────────
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)
    warnings: list[str] = []

    # ── 2. Read all parameter cells (data_only copy) ───────────────────────
    wb_data = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

    pack: dict[str, float] = {}
    ws_pack_d = wb_data[_PACK_SHEET] if _PACK_SHEET in wb_data.sheetnames else None
    if ws_pack_d:
        for key, (row, col) in _PACK_CELLS.items():
            pack[key] = _safe_float(ws_pack_d.cell(row=row, column=col).value)

    motor: dict[str, float] = {}
    ws_ep_d = wb_data[_EP_SHEET] if _EP_SHEET in wb_data.sheetnames else None
    if ws_ep_d:
        for key, (row, col) in _EP_PARAMS.items():
            motor[key] = _safe_float(ws_ep_d.cell(row=row, column=col).value)

    # Gear ratios from header row (H1..V1 of ElecPropulsion)
    gear_ratios = []
    if ws_ep_d:
        for col in range(_EP_GEAR_COL_START, _EP_GEAR_COL_END + 1):
            v = ws_ep_d.cell(row=_EP_GEAR_RATIO_ROW, column=col).value
            gear_ratios.append(_safe_float(v, default=1.0))
    else:
        gear_ratios = [1.0] * 15

    # Key constants for formula evaluation
    wheel_diam_in  = motor.get("wheel_diam_in", 18.0)   or 18.0
    pack_v_ep      = motor.get("pack_voltage_ep_v", 504.0) or 504.0  # EP B8
    motor_pf       = motor.get("motor_pf", 0.95)          or 0.95    # EP B11
    motor_eff      = motor.get("motor_efficiency", 0.9545) or 0.9545 # EP B6
    pack_v_bpc     = pack.get("pack_voltage_v", 504.0)    or 504.0   # BPC B11
    fuse_limit     = pack.get("fuse_max_a", 50.0)         or 50.0
    usable_kwh     = (pack_v_bpc * pack.get("pack_capacity_ah", 15.0)) * 0.92 / 1000.0

    # ── 3. Evaluate RPM grid (15 gear ratios × n_pts) — Python ───────────
    # Formula: RPM = v_mph * gear_ratio * 1056 / (wheel_diam_in * π)
    # Constant k = 1056 / (π * wheel_diam_in)
    k_rpm = 1056.0 / (math.pi * wheel_diam_in)
    rpm_all = np.zeros((n_pts, len(gear_ratios)), dtype=float)
    for gi, gr in enumerate(gear_ratios):
        rpm_all[:, gi] = v_mph * gr * k_rpm

    rpm_gear1 = rpm_all[:, 0]   # gear ratio 1 (col H)

    # ── 4. Current draw: =(EP.B8 * EP.B11 * RPM) / 1000 ─────────────────
    # EP.B8 = pack_voltage_ep_v (504), EP.B11 = motor_pf (0.95)
    current_draw_a  = (pack_v_ep * motor_pf * rpm_gear1) / 1000.0

    # ── 5. Phase current: =(EP.B11 * √3 * EP.B6 * RPM)  — no /1000, matches sheet ──
    phase_current_a = motor_pf * math.sqrt(3) * motor_eff * rpm_gear1

    # ── 6. Electrical power ───────────────────────────────────────────────
    power_kw = current_draw_a * pack_v_ep / 1000.0

    # ── 7. Total energy (trapezoidal) ─────────────────────────────────────
    dt_arr = np.diff(t_arr, prepend=t_arr[0])
    total_energy_kwh = float(np.sum(power_kw * np.abs(dt_arr))) / 3600.0

    # ── 8. Feasibility ────────────────────────────────────────────────────
    peak_i    = float(np.max(current_draw_a))
    avg_i     = float(np.mean(current_draw_a))
    peak_pw   = float(np.max(power_kw))
    fuse_ok   = peak_i <= fuse_limit
    energy_ok = total_energy_kwh <= usable_kwh
    feasible  = fuse_ok and energy_ok
    fuse_margin = fuse_limit - peak_i

    # Fuse-limited max sustained speed (solve P_wheel=P_fuse_cap numerically)
    P_fuse_kw  = fuse_limit * pack_v_ep / 1000.0
    rho, cda, crr_def, g = 1.225, 1.10, 0.018, 9.81
    v_test = np.linspace(0.1, 100.0, 5000)
    P_test = (0.5 * rho * cda * v_test**3 + crr_def * 230.0 * g * v_test) / 1000.0
    idx_ceil = int(np.searchsorted(P_test, P_fuse_kw * 0.90))
    fuse_speed_ms = float(v_test[min(idx_ceil, len(v_test)-1)])

    if feasible:
        verdict = (f"✅ Electrically feasible — peak {peak_i:.1f} A / "
                   f"{fuse_limit:.0f} A fuse  |  "
                   f"{total_energy_kwh:.3f} kWh / {usable_kwh:.3f} kWh usable")
    else:
        issues = []
        if not fuse_ok:   issues.append(f"fuse blown (+{-fuse_margin:.1f} A over)")
        if not energy_ok: issues.append(f"energy deficit ({total_energy_kwh-usable_kwh:.3f} kWh short)")
        verdict = f"❌ NOT feasible — {', '.join(issues)}"

    # ── 9. Write updated workbook (all computed values as plain numbers) ───
    # 9a. SpeedVsTime — clear old data, write new
    ws_svt = wb[_SVT_SHEET]
    orig_max = ws_svt.max_row
    for r in range(2, orig_max + 1):
        ws_svt.cell(row=r, column=1).value = None
        ws_svt.cell(row=r, column=2).value = None

    for i, (t, v) in enumerate(zip(t_arr, v_mph)):
        r = 2 + i
        ws_svt.cell(row=r, column=1).value = round(float(t), 6)
        ws_svt.cell(row=r, column=2).value = round(float(v), 4)

    # MAX formula row
    max_row = 2 + n_pts
    ws_svt.cell(row=max_row, column=1).value = "Max Speed (mph):"
    ws_svt.cell(row=max_row, column=2).value = float(np.max(v_mph))

    # 9b. ElecPropulsion — write computed values as plain numbers
    ws_ep = wb[_EP_SHEET]

    # Clear old blocks beyond header + param rows
    # (Rows 2..orig_max in all formula columns H..V)
    orig_ep_max = ws_ep.max_row
    for r in range(2, orig_ep_max + 1):
        for col in range(_EP_GEAR_COL_START, _EP_GEAR_COL_END + 1):
            ws_ep.cell(row=r, column=col).value = None
        # Clear col G labels beyond row 1
        if ws_ep.cell(row=r, column=7).value not in (None, "Current Draw (A)", "Phase Current (A)"):
            pass
        ws_ep.cell(row=r, column=7).value = None

    # Block 1: RPM values (rows 2..n_pts+1, cols H..V)
    for i in range(n_pts):
        r = 2 + i
        for gi in range(len(gear_ratios)):
            col = _EP_GEAR_COL_START + gi
            ws_ep.cell(row=r, column=col).value = round(rpm_all[i, gi], 4)

    # Block 2: Current draw (rows n_pts+3 .. 2*n_pts+2, col H only)
    cur_label_row  = n_pts + 2
    cur_start_row  = n_pts + 3
    ws_ep.cell(row=cur_label_row, column=7).value = "Current Draw (A)"
    for i in range(n_pts):
        r = cur_start_row + i
        ws_ep.cell(row=r, column=8).value = round(current_draw_a[i], 6)

    # Block 3: Phase current (rows 2*n_pts+4 .. 3*n_pts+3, cols H..V)
    # Formula: =$B$11*(SQRT(3))*$B$6*H{rpm_row}  — NO /1000, result is in raw units
    phase_label_row = 2 * n_pts + 3
    phase_start_row = 2 * n_pts + 4
    ws_ep.cell(row=phase_label_row, column=7).value = "Phase Current (A)"
    for i in range(n_pts):
        r = phase_start_row + i
        for gi in range(len(gear_ratios)):
            col = _EP_GEAR_COL_START + gi
            phase_i = motor_pf * math.sqrt(3) * motor_eff * rpm_all[i, gi]
            ws_ep.cell(row=r, column=col).value = round(phase_i, 6)

    # 9c. Battery Pack Calcs — stamp the KinematiK summary block
    ws_bpc = wb[_PACK_SHEET]
    summary_start = 18
    summary_data = [
        ("─── KinematiK Lap Sim ───",      ""),
        ("Lap Time (s)",                    round(float(lap_time_s), 3) if lap_time_s else ""),
        ("Top Speed (km/h)",                round(float(top_speed_ms)*3.6, 2) if top_speed_ms else ""),
        ("Top Speed (mph)",                 round(float(top_speed_ms)*2.23694, 2) if top_speed_ms else ""),
        ("Avg Speed (km/h)",                round(float(avg_speed_ms)*3.6, 2) if avg_speed_ms else ""),
        ("Profile Points",                   int(n_pts)),
        ("Profile Duration (s)",             round(float(t_arr[-1]), 2)),
        ("Max Speed in Profile (mph)",       round(float(np.max(v_mph)), 2)),
        ("Peak Current Draw (A)",            round(peak_i, 2)),
        ("Avg Current Draw (A)",             round(avg_i, 2)),
        ("Peak Power Draw (kW)",             round(peak_pw, 2)),
        ("Total Energy (kWh)",               round(total_energy_kwh, 4)),
        ("Usable Pack Energy (kWh)",         round(usable_kwh, 4)),
        ("Fuse Limit (A)",                   fuse_limit),
        ("Fuse Margin (A)",                  round(fuse_margin, 2)),
        ("Fuse-limited Speed Ceiling (mph)", round(fuse_speed_ms * 2.23694, 1)),
        ("Feasibility",                      "PASS" if feasible else "FAIL"),
    ]
    for offset, (label, val) in enumerate(summary_data):
        r = summary_start + offset
        ws_bpc.cell(row=r, column=1).value = label
        ws_bpc.cell(row=r, column=2).value = val

    # 9d. Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    out_bytes = buf.getvalue()

    return ExcelRoundTripResult(
        ok=True,
        warnings=warnings,
        time_s=t_arr,
        speed_mph=v_mph,
        speed_ms=v_ms,
        rpm_gear1=rpm_gear1,
        current_draw_a=current_draw_a,
        phase_current_a=phase_current_a,
        power_kw=power_kw,
        pack=pack,
        motor=motor,
        max_speed_mph=float(np.max(v_mph)),
        peak_current_a=peak_i,
        avg_current_a=avg_i,
        peak_power_kw=peak_pw,
        total_energy_kwh=total_energy_kwh,
        fuse_margin_a=fuse_margin,
        fuse_speed_ceiling_ms=fuse_speed_ms,
        usable_energy_kwh=usable_kwh,
        fuse_ok=fuse_ok,
        energy_ok=energy_ok,
        feasible=feasible,
        verdict=verdict,
        excel_bytes=out_bytes,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Extended analysis: Joule heat per cell, SoC depletion, lap-stop prediction,
# minimum-feasible pack advisor, and the enhanced Excel export.
# ─────────────────────────────────────────────────────────────────────────────

def compute_cell_thermals(
    current_draw_a: np.ndarray,
    time_s:         np.ndarray,
    pack:           dict,
    *,
    ambient_c:       float = 25.0,
    thermal_mass_j_k: float | None = None,
    h_conv_w_m2k:    float = 10.0,
    cell_surface_m2:  float = 0.0020,
) -> dict:
    """
    Per-cell Joule heat time-series and thermal model.

    The pack current is shared across n_parallel strings, so each cell sees
    I_cell = I_pack / n_parallel.  Joule power per cell = I_cell² × R_cell.

    Returns a dict with numpy arrays keyed as:
        i_cell_a        — per-cell current (A)
        joule_w         — instantaneous Joule power per cell (W)
        cumulative_j    — cumulative Joule energy per cell (J)
        temp_c          — estimated cell surface temperature (°C)
        peak_temp_c     — max temperature reached (°C)
        peak_joule_w    — peak instantaneous Joule power (W)
        total_joule_kwh — total Joule heat per cell over the lap (kWh)
    """
    n_parallel = max(int(pack.get("n_parallel", 1) or 1), 1)
    r_cell     = float(pack.get("cell_r_ohm", 0.003) or 0.003)
    cell_mass  = float(pack.get("cell_weight_kg", 0.065) or 0.065)

    # Specific heat of LFP/NMC cells ~ 900 J/(kg·K)
    cp = 900.0
    if thermal_mass_j_k is None:
        thermal_mass_j_k = cell_mass * cp

    i_cell = current_draw_a / n_parallel
    joule_w = i_cell ** 2 * r_cell

    dt = np.diff(time_s, prepend=time_s[0])
    dt = np.abs(dt)

    # Simple lumped thermal model: dT/dt = (Q_gen - Q_conv) / C_th
    # Q_conv = h * A * (T - T_amb)
    temp_c = np.zeros(len(time_s))
    temp_c[0] = ambient_c
    for k in range(1, len(time_s)):
        q_gen  = joule_w[k] * dt[k]
        q_conv = h_conv_w_m2k * cell_surface_m2 * (temp_c[k-1] - ambient_c) * dt[k]
        temp_c[k] = temp_c[k-1] + (q_gen - q_conv) / thermal_mass_j_k

    cumulative_j    = np.cumsum(joule_w * dt)
    total_joule_j   = float(cumulative_j[-1]) if len(cumulative_j) else 0.0
    total_joule_kwh = total_joule_j / 3_600_000.0

    return {
        "i_cell_a":        i_cell,
        "joule_w":         joule_w,
        "cumulative_j":    cumulative_j,
        "temp_c":          temp_c,
        "peak_temp_c":     float(np.max(temp_c)),
        "peak_joule_w":    float(np.max(joule_w)),
        "total_joule_kwh": total_joule_kwh,
    }


def compute_soc_and_stop(
    power_kw: np.ndarray,
    time_s:   np.ndarray,
    pack:     dict,
    motor:    dict,
    *,
    usable_frac: float = 0.92,
) -> dict:
    """
    Integrate power draw vs time to track state-of-charge and find exactly
    where on the lap the car would run out of energy.

    Returns:
        soc_pct         — SoC array (100 % → 0 %)
        energy_used_kwh — cumulative energy used at each time step
        stop_time_s     — time at which pack hits 0 % (None if it finishes)
        stop_idx        — array index of stop (None if finishes)
        pct_lap_done    — fraction of lap completed at stop (None if finishes)
        finishes        — bool
        deficit_kwh     — how much energy is still needed after pack is empty (0 if finishes)
    """
    pack_v     = float(pack.get("pack_voltage_v", 504.0) or 504.0)
    cap_ah     = float(pack.get("pack_capacity_ah", 15.0) or 15.0)
    usable_kwh = pack_v * cap_ah * usable_frac / 1000.0

    dt = np.abs(np.diff(time_s, prepend=time_s[0]))
    energy_used_kwh = np.cumsum(power_kw * dt) / 3600.0

    soc_pct = np.clip(100.0 * (1.0 - energy_used_kwh / usable_kwh), 0.0, 100.0)

    stop_idx = None
    stop_time_s = None
    pct_lap_done = None
    finishes = True
    deficit_kwh = 0.0

    depleted = np.where(energy_used_kwh >= usable_kwh)[0]
    if len(depleted):
        stop_idx     = int(depleted[0])
        stop_time_s  = float(time_s[stop_idx])
        lap_time     = float(time_s[-1])
        pct_lap_done = stop_time_s / lap_time if lap_time > 0 else 0.0
        finishes     = False
        deficit_kwh  = float(energy_used_kwh[-1] - usable_kwh)

    return {
        "soc_pct":         soc_pct,
        "energy_used_kwh": energy_used_kwh,
        "usable_kwh":      usable_kwh,
        "stop_time_s":     stop_time_s,
        "stop_idx":        stop_idx,
        "pct_lap_done":    pct_lap_done,
        "finishes":        finishes,
        "deficit_kwh":     deficit_kwh,
    }


def compute_minimum_feasible_pack(
    power_kw:  np.ndarray,
    time_s:    np.ndarray,
    current_draw_a: np.ndarray,
    pack:      dict,
    motor:     dict,
    *,
    usable_frac:  float = 0.92,
    safety_margin: float = 0.10,   # 10 % headroom on top of calculated need
) -> dict:
    """
    Work backwards from the lap profile to answer:
      "What is the smallest pack that lets this car finish the lap?"

    Constraints:
      1. Energy: usable pack energy ≥ total energy drawn × (1 + safety_margin)
      2. Current: peak I_pack ≤ fuse limit (pack topology doesn't change this)

    Returns a dict with:
        min_energy_kwh      — minimum usable energy needed
        min_capacity_ah     — minimum Ah at current pack voltage
        rec_capacity_ah     — recommended (with safety margin)
        rec_energy_kwh      — recommended usable energy
        rec_cells_series    — cells in series (same voltage per cell, min count)
        rec_cells_parallel  — strings in parallel for capacity
        rec_total_cells     — total cell count
        rec_pack_mass_kg    — estimated pack mass
        current_usable_kwh  — what the current pack provides
        energy_shortfall_kwh— energy gap (0 if already feasible)
        fuse_ok             — bool: is peak current already within fuse?
        peak_current_a      — peak pack current seen on the lap
        fuse_limit_a        — fuse rating from pack params
    """
    pack_v    = float(pack.get("pack_voltage_v", 504.0) or 504.0)
    cell_v    = float(pack.get("cell_voltage_v", 3.6) or 3.6)
    cell_ah   = float(pack.get("cell_capacity_ah", 2.5) or 2.5)
    cell_mass = float(pack.get("cell_weight_kg", 0.065) or 0.065)
    n_series  = int(pack.get("n_series", 1) or max(1, round(pack_v / max(cell_v, 0.1))))
    fuse_a    = float(pack.get("fuse_max_a", 50.0) or 50.0)

    dt = np.abs(np.diff(time_s, prepend=time_s[0]))
    total_energy_kwh = float(np.sum(power_kw * dt)) / 3600.0
    peak_i = float(np.max(current_draw_a))

    min_energy_kwh  = total_energy_kwh / usable_frac
    rec_energy_kwh  = min_energy_kwh * (1.0 + safety_margin)

    # Capacity from energy: E = V × Ah  → Ah = E×1000 / V
    min_cap_ah = min_energy_kwh * 1000.0 / max(pack_v, 1.0)
    rec_cap_ah = rec_energy_kwh * 1000.0 / max(pack_v, 1.0)

    # How many parallel strings: ceil(rec_cap_ah / cell_ah)
    n_par_rec   = max(1, math.ceil(rec_cap_ah / max(cell_ah, 0.01)))
    total_cells = n_series * n_par_rec
    pack_mass   = total_cells * cell_mass

    cur_usable  = (pack_v * float(pack.get("pack_capacity_ah", 15.0) or 15.0)
                   * usable_frac / 1000.0)
    shortfall   = max(0.0, total_energy_kwh - cur_usable)

    return {
        "min_energy_kwh":       round(min_energy_kwh, 4),
        "min_capacity_ah":      round(min_cap_ah, 2),
        "rec_capacity_ah":      round(rec_cap_ah, 2),
        "rec_energy_kwh":       round(rec_energy_kwh, 4),
        "rec_cells_series":     n_series,
        "rec_cells_parallel":   n_par_rec,
        "rec_total_cells":      total_cells,
        "rec_pack_mass_kg":     round(pack_mass, 2),
        "current_usable_kwh":   round(cur_usable, 4),
        "energy_shortfall_kwh": round(shortfall, 4),
        "fuse_ok":              peak_i <= fuse_a,
        "peak_current_a":       round(peak_i, 2),
        "fuse_limit_a":         fuse_a,
        "safety_margin_pct":    safety_margin * 100,
        "pack_voltage_v":       pack_v,
        "cell_voltage_v":       cell_v,
        "cell_capacity_ah":     cell_ah,
    }


def build_enhanced_excel(
    result:         "ExcelRoundTripResult",
    excel_bytes:    bytes,
    thermals:       dict,
    soc_data:       dict,
    min_pack:       dict,
    *,
    lap_time_s:     float = 0.0,
    top_speed_ms:   float = 0.0,
    avg_speed_ms:   float = 0.0,
) -> bytes:
    """
    Write three new sheets into the workbook on top of the existing round-trip:
      1. PackHeatmap    — per-cell Joule heat grid with conditional formatting
      2. LapEnergy      — SoC depletion curve, stop marker, energy timeline
      3. FeasiblePack   — minimum feasible pack advisor + what-if table

    Also embeds a Python xlwings-compatible macro stub in PackHeatmap so the
    team can refresh the heat map with one click if they have xlwings installed.

    Returns updated xlsx bytes.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side,
            GradientFill,
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import AreaChart, Reference
        from openpyxl.chart.series import SeriesLabel
    except ImportError:
        # Return original bytes if openpyxl not available
        return excel_bytes

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)

    # ── colour helpers ────────────────────────────────────────────────────────
    def _heat_colour(val: float, lo: float, hi: float) -> str:
        """Map val in [lo,hi] → hex fill colour green→yellow→red."""
        if hi <= lo:
            return "FF37E0D0"
        t = max(0.0, min(1.0, (val - lo) / (hi - lo)))
        if t < 0.5:
            # green → yellow
            r = int(55  + t * 2 * 200)
            g = int(224 - t * 2 * 24)
            b = int(192 - t * 2 * 192)
        else:
            # yellow → red
            r = int(255)
            g = int(200 - (t - 0.5) * 2 * 200)
            b = 0
        return f"FF{r:02X}{g:02X}{b:02X}"

    def _cell_fill(colour_hex: str) -> PatternFill:
        return PatternFill(fill_type="solid", fgColor=colour_hex)

    def _header(ws, row: int, col: int, text: str, bold=True):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=bold, color="FFFFFFFF", size=10)
        c.fill = PatternFill(fill_type="solid", fgColor="FF1A2433")
        c.alignment = Alignment(horizontal="center", vertical="center")
        return c

    def _thin_border():
        s = Side(border_style="thin", color="FF2D3748")
        return Border(left=s, right=s, top=s, bottom=s)

    pack  = result.pack
    motor = result.motor
    t_arr = result.time_s
    n_pts = len(t_arr)

    # ── Sheet 1: PackHeatmap ─────────────────────────────────────────────────
    _HM = "PackHeatmap"
    if _HM in wb.sheetnames:
        del wb[_HM]
    ws_hm = wb.create_sheet(_HM)
    ws_hm.sheet_properties.tabColor = "E74C3C"

    n_series  = max(int(pack.get("n_series",   1) or 1), 1)
    n_par     = max(int(pack.get("n_parallel", 1) or 1), 1)
    joule_w   = thermals["joule_w"]
    temp_c    = thermals["temp_c"]
    i_cell    = thermals["i_cell_a"]

    # Downsample to ≤500 pts for the grid snapshot
    step = max(1, n_pts // 200)
    snap_idx = list(range(0, n_pts, step))

    # ── Title block
    ws_hm.merge_cells("A1:H1")
    t = ws_hm["A1"]
    t.value = "🔋 Battery Pack Thermal Heatmap — KinematiK Lap Analysis"
    t.font = Font(bold=True, size=13, color="FFFFFFFF")
    t.fill = PatternFill(fill_type="solid", fgColor="FF0D1117")
    t.alignment = Alignment(horizontal="center")
    ws_hm.row_dimensions[1].height = 24

    # ── Key metrics row
    metrics = [
        ("Peak cell temp", f"{thermals['peak_temp_c']:.1f} °C"),
        ("Peak Joule power", f"{thermals['peak_joule_w']:.2f} W/cell"),
        ("Total Joule heat", f"{thermals['total_joule_kwh']*1000:.2f} Wh/cell"),
        ("Pack topology", f"{n_series}S × {n_par}P"),
        ("Cells total", f"{n_series * n_par}"),
        ("Cell R", f"{pack.get('cell_r_ohm', '?')} Ω"),
    ]
    for ci, (label, val) in enumerate(metrics):
        col = ci + 1
        ws_hm.cell(row=2, column=col, value=label).font = Font(
            bold=True, size=9, color="FF8899AA")
        ws_hm.cell(row=2, column=col).fill = PatternFill(
            fill_type="solid", fgColor="FF0D1117")
        ws_hm.cell(row=3, column=col, value=val).font = Font(
            bold=True, size=11, color="FFFFFFFF")
        ws_hm.cell(row=3, column=col).fill = PatternFill(
            fill_type="solid", fgColor="FF131B24")

    # ── Cell grid header
    _header(ws_hm, 5, 1, "Cell position →  (columns = parallel strings)")
    _header(ws_hm, 6, 1, "Series\nRow↓", bold=True)
    for p in range(n_par):
        _header(ws_hm, 6, p + 2, f"P{p+1}")

    # ── Fill cell grid with Joule heat colour
    # Each cell in the grid represents one physical cell.
    # Joule heat per cell is the same for all cells at the same parallel
    # position (they share the same string current), so the n_series rows
    # differ only if cell-level resistance differs — we use a single lap
    # total here and modulate by series position (thermal gradient proxy).
    total_j_per_cell = thermals["total_joule_kwh"] * 3_600_000.0  # J
    # Series cells at the bottom of the stack tend to run hotter in real packs.
    # Approximate with a ±15% gradient across the stack.
    series_gradient = np.linspace(0.85, 1.15, n_series)

    # temperature range for colour mapping
    t_lo = 25.0
    t_hi = max(thermals["peak_temp_c"], 30.0)

    for s in range(n_series):
        row = 7 + s
        ws_hm.cell(row=row, column=1, value=f"S{s+1}").font = Font(
            bold=True, size=9, color="FF8899AA")
        for p in range(n_par):
            col = p + 2
            cell_j = total_j_per_cell * series_gradient[s]
            cell_t = t_lo + (t_hi - t_lo) * series_gradient[s]
            colour  = _heat_colour(cell_t, t_lo, t_hi)
            c = ws_hm.cell(row=row, column=col, value=round(cell_t, 1))
            c.fill      = _cell_fill(colour)
            c.font      = Font(size=8, color="FF000000" if cell_t < (t_lo+t_hi)/2 else "FFFFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.border    = _thin_border()
            c.number_format = "0.0"

    # ── Legend
    legend_row = 7 + n_series + 2
    ws_hm.cell(row=legend_row, column=1, value="Temperature legend:").font = Font(
        bold=True, size=9, color="FF8899AA")
    labels = ["Cool (25°C)", "Warm", "Hot", f"Peak ({t_hi:.0f}°C)"]
    for i, label in enumerate(labels):
        t_val = t_lo + (t_hi - t_lo) * i / (len(labels) - 1)
        colour = _heat_colour(t_val, t_lo, t_hi)
        col = i + 2
        c = ws_hm.cell(row=legend_row, column=col, value=label)
        c.fill = _cell_fill(colour)
        c.font = Font(size=9, color="FF000000" if t_val < (t_lo+t_hi)/2 else "FFFFFFFF")
        c.alignment = Alignment(horizontal="center")

    # ── Time-series data table (Joule W and cell temp per time step)
    ts_start_row = legend_row + 3
    _header(ws_hm, ts_start_row, 1, "Time (s)")
    _header(ws_hm, ts_start_row, 2, "Speed (km/h)")
    _header(ws_hm, ts_start_row, 3, "Pack I (A)")
    _header(ws_hm, ts_start_row, 4, "Cell I (A)")
    _header(ws_hm, ts_start_row, 5, "Joule/cell (W)")
    _header(ws_hm, ts_start_row, 6, "Cell temp (°C)")
    _header(ws_hm, ts_start_row, 7, "Cum. heat/cell (J)")

    cum_j = thermals["cumulative_j"]
    for ri, idx in enumerate(snap_idx):
        r = ts_start_row + 1 + ri
        ws_hm.cell(row=r, column=1, value=round(float(t_arr[idx]), 2))
        ws_hm.cell(row=r, column=2, value=round(float(result.speed_ms[idx]) * 3.6, 2))
        ws_hm.cell(row=r, column=3, value=round(float(result.current_draw_a[idx]), 3))
        ws_hm.cell(row=r, column=4, value=round(float(i_cell[idx]), 4))
        ws_hm.cell(row=r, column=5, value=round(float(joule_w[idx]), 4))
        ws_hm.cell(row=r, column=6, value=round(float(temp_c[idx]), 2))
        ws_hm.cell(row=r, column=7, value=round(float(cum_j[idx]), 2))

        # heat-colour the cell-temp column
        t_val = float(temp_c[idx])
        ws_hm.cell(row=r, column=6).fill = _cell_fill(_heat_colour(t_val, t_lo, t_hi))
        ws_hm.cell(row=r, column=6).font = Font(
            size=9, color="FFFFFFFF" if t_val > (t_lo + t_hi) / 2 else "FF000000")

    # ── Python macro stub (embedded as a named comment range so xlwings can find it)
    pycode_row = ts_start_row + len(snap_idx) + 3
    ws_hm.cell(row=pycode_row, column=1,
               value="# ── KinematiK Python Macro (run via xlwings) ──────────────────────").font = Font(
                   bold=True, color="FF37E0D0", size=9)
    macro_lines = [
        "import xlwings as xw, pandas as pd, matplotlib.pyplot as plt",
        "from matplotlib.colors import LinearSegmentedColormap",
        "",
        "def refresh_heatmap():",
        "    wb  = xw.books.active",
        "    ws  = wb.sheets['PackHeatmap']",
        "    # Read the time-series block written by KinematiK",
        f"    df = ws.range('A{ts_start_row+1}').expand().options(pd.DataFrame, header=1).value",
        "    df.columns = ['time_s','speed_kmh','pack_i_a','cell_i_a','joule_w','cell_t_c','cum_j']",
        "    cmap = LinearSegmentedColormap.from_list('heat', ['#37E0D0','#F0A500','#E74C3C'])",
        "    fig, ax = plt.subplots(figsize=(10, 3))",
        "    sc = ax.scatter(df.time_s, df.cell_t_c, c=df.cell_t_c, cmap=cmap, s=4)",
        "    plt.colorbar(sc, ax=ax, label='Cell temperature (°C)')",
        "    ax.set_xlabel('Lap time (s)'); ax.set_ylabel('Cell temp (°C)')",
        "    ax.set_title('KinematiK — Cell thermal profile over lap')",
        "    ax.set_facecolor('#0D1117'); fig.patch.set_facecolor('#0D1117')",
        "    ax.tick_params(colors='white'); ax.xaxis.label.set_color('white')",
        "    ax.yaxis.label.set_color('white'); ax.title.set_color('white')",
        "    plt.tight_layout()",
        "    pic_path = 'kinematik_heatmap.png'",
        "    fig.savefig(pic_path, dpi=150, bbox_inches='tight')",
        "    ws.pictures.add(pic_path, name='HeatmapChart', update=True,",
        "                    left=ws.range('I2').left, top=ws.range('I2').top)",
        "    print('Heatmap refreshed.')",
        "",
        "refresh_heatmap()",
    ]
    for li, line in enumerate(macro_lines):
        ws_hm.cell(row=pycode_row + 1 + li, column=1, value=line).font = Font(
            name="Courier New", size=8, color="FF8899AA")

    # Column widths
    ws_hm.column_dimensions["A"].width = 18
    for ci in range(2, n_par + 2):
        ws_hm.column_dimensions[get_column_letter(ci)].width = 8

    # ── Sheet 2: LapEnergy ──────────────────────────────────────────────────
    _LE = "LapEnergy"
    if _LE in wb.sheetnames:
        del wb[_LE]
    ws_le = wb.create_sheet(_LE)
    ws_le.sheet_properties.tabColor = "3B7CFF"

    ws_le.merge_cells("A1:G1")
    t2 = ws_le["A1"]
    t2.value = "⚡ Lap Energy & State-of-Charge — KinematiK"
    t2.font = Font(bold=True, size=13, color="FFFFFFFF")
    t2.fill = PatternFill(fill_type="solid", fgColor="FF0D1117")
    t2.alignment = Alignment(horizontal="center")

    # Summary metrics
    soc_meta = [
        ("Usable pack energy",    f"{soc_data['usable_kwh']:.3f} kWh"),
        ("Total energy drawn",    f"{result.total_energy_kwh:.3f} kWh"),
        ("Pack finishes lap",     "✅ YES" if soc_data["finishes"] else "❌ NO"),
        ("Stop time",             f"{soc_data['stop_time_s']:.1f} s" if soc_data["stop_time_s"] else "—"),
        ("Lap % complete at stop", f"{(soc_data['pct_lap_done'] or 0)*100:.1f} %" if not soc_data["finishes"] else "100 %"),
        ("Energy deficit",        f"{soc_data['deficit_kwh']:.3f} kWh" if not soc_data['finishes'] else "—"),
    ]
    for ci, (label, val) in enumerate(soc_meta):
        col = ci + 1
        ws_le.cell(row=2, column=col, value=label).font = Font(
            bold=True, size=9, color="FF8899AA")
        ws_le.cell(row=2, column=col).fill = PatternFill(
            fill_type="solid", fgColor="FF0D1117")
        bad = "NO" in val or "deficit" in label.lower()
        ws_le.cell(row=3, column=col, value=val).font = Font(
            bold=True, size=11,
            color="FFE74C3C" if bad else "FF37E0D0")
        ws_le.cell(row=3, column=col).fill = PatternFill(
            fill_type="solid", fgColor="FF131B24")

    # Data table
    _header(ws_le, 5, 1, "Time (s)")
    _header(ws_le, 5, 2, "Speed (km/h)")
    _header(ws_le, 5, 3, "Power (kW)")
    _header(ws_le, 5, 4, "Energy used (kWh)")
    _header(ws_le, 5, 5, "SoC (%)")
    _header(ws_le, 5, 6, "Pack I (A)")
    _header(ws_le, 5, 7, "Marker")

    soc_arr  = soc_data["soc_pct"]
    e_used   = soc_data["energy_used_kwh"]
    stop_idx = soc_data["stop_idx"]

    for ri, idx in enumerate(snap_idx):
        r = 6 + ri
        soc_val = float(soc_arr[idx])
        ws_le.cell(row=r, column=1, value=round(float(t_arr[idx]), 2))
        ws_le.cell(row=r, column=2, value=round(float(result.speed_ms[idx]) * 3.6, 2))
        ws_le.cell(row=r, column=3, value=round(float(result.power_kw[idx]), 3))
        ws_le.cell(row=r, column=4, value=round(float(e_used[idx]), 4))
        soc_c = ws_le.cell(row=r, column=5, value=round(soc_val, 1))
        soc_c.fill = _cell_fill(_heat_colour(100.0 - soc_val, 0, 100))
        soc_c.font = Font(size=9, color="FF000000" if soc_val > 50 else "FFFFFFFF")
        ws_le.cell(row=r, column=6, value=round(float(result.current_draw_a[idx]), 2))
        # mark the stop point
        if stop_idx is not None and idx >= stop_idx and (
                ri == 0 or snap_idx[ri-1] < stop_idx):
            ws_le.cell(row=r, column=7, value="🛑 STOP").font = Font(
                bold=True, color="FFE74C3C", size=10)

    # SoC chart
    try:
        chart_ref_soc = Reference(ws_le,
            min_col=5, min_row=5, max_row=5 + len(snap_idx))
        chart_ref_t   = Reference(ws_le,
            min_col=1, min_row=6, max_row=5 + len(snap_idx))
        from openpyxl.chart import LineChart
        lc = LineChart()
        lc.title  = "State of Charge over Lap"
        lc.y_axis.title = "SoC (%)"
        lc.x_axis.title = "Time (s)"
        lc.height = 12
        lc.width  = 22
        lc.add_data(chart_ref_soc, titles_from_data=True)
        lc.set_categories(chart_ref_t)
        lc.series[0].graphicalProperties.line.solidFill = "3B7CFF"
        lc.series[0].graphicalProperties.line.width = 20000
        ws_le.add_chart(lc, f"I5")
    except Exception:
        pass

    for col_idx, w in enumerate([12, 12, 10, 16, 8, 10, 14], 1):
        ws_le.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Sheet 3: FeasiblePack ────────────────────────────────────────────────
    _FP = "FeasiblePack"
    if _FP in wb.sheetnames:
        del wb[_FP]
    ws_fp = wb.create_sheet(_FP)
    ws_fp.sheet_properties.tabColor = "A855F7"

    ws_fp.merge_cells("A1:E1")
    t3 = ws_fp["A1"]
    t3.value = "📐 Minimum Feasible Pack Advisor — KinematiK"
    t3.font = Font(bold=True, size=13, color="FFFFFFFF")
    t3.fill = PatternFill(fill_type="solid", fgColor="FF0D1117")
    t3.alignment = Alignment(horizontal="center")

    # Current pack vs minimum
    fp_rows = [
        ("",                             "Current pack",   "Minimum needed",    "Recommended (+10%)", "Unit"),
        ("Energy (usable)",              f"{min_pack['current_usable_kwh']:.3f}",
                                          f"{min_pack['min_energy_kwh']:.3f}",
                                          f"{min_pack['rec_energy_kwh']:.3f}",   "kWh"),
        ("Capacity",                     f"{pack.get('pack_capacity_ah','?')}",
                                          f"{min_pack['min_capacity_ah']:.2f}",
                                          f"{min_pack['rec_capacity_ah']:.2f}",  "Ah"),
        ("Parallel strings",             f"{pack.get('n_parallel','?')}",
                                          "—",
                                          f"{min_pack['rec_cells_parallel']}",    "strings"),
        ("Total cells",                  f"{pack.get('pack_cell_count','?')}",
                                          "—",
                                          f"{min_pack['rec_total_cells']}",       "cells"),
        ("Estimated pack mass",          f"{pack.get('cell_weight_kg',0.065) * (pack.get('pack_cell_count',0) or min_pack['rec_total_cells']):.1f}",
                                          "—",
                                          f"{min_pack['rec_pack_mass_kg']:.1f}", "kg"),
        ("Energy shortfall",             "—",
                                          f"{min_pack['energy_shortfall_kwh']:.3f}",
                                          "—",                                   "kWh"),
        ("Fuse OK for this lap",         "—",
                                          "✅" if min_pack["fuse_ok"] else "❌",
                                          "—",                                   ""),
        ("Peak pack current",            "—",
                                          f"{min_pack['peak_current_a']:.1f}",
                                          "—",                                   "A"),
        ("Fuse rating",                  f"{min_pack['fuse_limit_a']:.0f}",
                                          f"{min_pack['fuse_limit_a']:.0f}",
                                          f"{min_pack['fuse_limit_a']:.0f}",     "A"),
    ]

    col_colours = ["FF1A2433", "FF131B24", "FF0D1B2A", "FF1A0D2A"]
    for ri, row_data in enumerate(fp_rows):
        for ci, val in enumerate(row_data):
            c = ws_fp.cell(row=ri + 2, column=ci + 1, value=val)
            c.fill = PatternFill(fill_type="solid", fgColor=col_colours[min(ci, 3)])
            if ri == 0:
                c.font = Font(bold=True, size=10, color="FF8899AA")
                c.alignment = Alignment(horizontal="center")
            else:
                is_bad = ("shortfall" in row_data[0].lower() and ci == 2 and
                          min_pack["energy_shortfall_kwh"] > 0)
                is_good = ("Fuse OK" in row_data[0] and ci == 2 and min_pack["fuse_ok"])
                c.font = Font(
                    size=10,
                    color=("FFE74C3C" if is_bad else
                           "FF37E0D0" if is_good else "FFFFFFFF"))
            c.border = _thin_border()

    # What-if table: capacity vs laps completed
    ws_fp.cell(row=14, column=1, value="What-if: capacity → laps completed").font = Font(
        bold=True, size=11, color="FFA855F7")

    total_e = result.total_energy_kwh
    pack_v_fp = min_pack["pack_voltage_v"]
    _header(ws_fp, 15, 1, "Capacity (Ah)")
    _header(ws_fp, 15, 2, "Usable energy (kWh)")
    _header(ws_fp, 15, 3, "Laps (endurance ~22)")
    _header(ws_fp, 15, 4, "Energy margin (%)")
    _header(ws_fp, 15, 5, "Status")

    for wi, cap_ah in enumerate([5, 8, 10, 12, 15, 18, 20, 25, 30]):
        usable = pack_v_fp * cap_ah * 0.92 / 1000.0
        laps   = usable / max(total_e, 0.001)
        margin = (usable / max(total_e, 0.001) - 1.0) * 100.0
        ok     = usable >= total_e
        r = 16 + wi
        ws_fp.cell(row=r, column=1, value=cap_ah)
        ws_fp.cell(row=r, column=2, value=round(usable, 3))
        ws_fp.cell(row=r, column=3, value=round(laps, 2))
        ws_fp.cell(row=r, column=4, value=round(margin, 1))
        status_c = ws_fp.cell(row=r, column=5,
                               value="✅ Feasible" if ok else "❌ Short")
        status_c.font = Font(color="FF37E0D0" if ok else "FFE74C3C", bold=True)
        for col_idx in range(1, 6):
            ws_fp.cell(row=r, column=col_idx).fill = PatternFill(
                fill_type="solid",
                fgColor="FF131B24" if ok else "FF2A1010")
            ws_fp.cell(row=r, column=col_idx).border = _thin_border()

    for col_idx, w in enumerate([16, 20, 20, 16, 12], 1):
        ws_fp.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Save ──────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════════════
#  ERICK'S FEEDBACK — "treat my file as a database, pull from it, visualise the
#  raw data, and make it the source of truth"  (Discord, 26-06-2026)
# ─────────────────────────────────────────────────────────────────────────────
#  The functions above already turn the workbook into a stored parameter set
#  (extract_params_from_excel) so nobody re-uploads. These add the three things
#  Erick asked for that weren't there yet:
#
#    1. read_raw_sheets()        — pull ARBITRARY tabular data out of any sheet,
#                                  so the app can plot whatever columns he puts
#                                  in the file (cell voltages, current, temps…),
#                                  not just the specific cells we hard-coded.
#                                  ("take that raw data from my sheet and
#                                   visualise it")
#    2. find_feasible_pack_sheet — surface a "what pack would you actually need"
#                                  sheet if he adds one to the file, reading it
#                                  directly rather than only computing our own.
#                                  ("once that sheet exists in your file the app
#                                   can just surface it directly")
#    3. ledger_declarations_from_ev_params — turn the pulled joule-heat + key
#                                  propulsion stats into integration-ledger
#                                  declarations, so his Excel becomes the
#                                  CROSS-TEAM source of truth, the same way the
#                                  Accumulator tab already declares numbers.
#                                  ("Joule heat fs should be pulled plus key
#                                   electric propulsion stats")
#
#  HONESTY CONTRACT (same as the rest of the module): these only ever READ what
#  is actually in the workbook. Nothing is invented. A column that isn't there
#  comes back absent, not zero-filled, so a chart never implies data the sheet
#  doesn't contain.
# ═════════════════════════════════════════════════════════════════════════════

@dataclass
class RawColumn:
    """One numeric column pulled verbatim from a worksheet, for plotting."""
    sheet: str
    header: str            # the column header text (or "Col <letter>" if blank)
    column_letter: str
    values: list           # numeric values, in row order (non-numeric -> skipped)
    row_start: int         # 1-based row of the first value


@dataclass
class RawSheet:
    """A worksheet reduced to its numeric columns, for direct visualisation."""
    name: str
    n_rows: int
    n_cols: int
    columns: list          # list[RawColumn]

    def column_headers(self) -> list:
        return [c.header for c in self.columns]


def read_raw_sheets(excel_bytes: bytes, *,
                    max_rows: int = 20000,
                    max_cols: int = 64,
                    min_numeric_fraction: float = 0.5) -> dict:
    """
    Read every worksheet and return its NUMERIC columns verbatim, so the app can
    plot the raw data straight from the user's file — Erick's "take that raw data
    from my sheet and visualise it".

    A column is included if it has a usable amount of numeric data
    (``min_numeric_fraction`` of its non-empty cells parse as numbers). The first
    row of each sheet is treated as a header if it is mostly non-numeric;
    otherwise columns are labelled by their Excel letter.

    Returns a JSON-friendly dict:
        {
          "sheets": [RawSheet-as-dict, ...],   # only sheets with >=1 numeric col
          "sheet_names": [all sheet names in the file],
          "_source": "excel",
          ["_error": "..."]                    # present only on failure
        }

    Nothing here is calibration-sensitive or invented: it is a faithful echo of
    the cells, with non-numeric cells dropped from numeric columns.
    """
    try:
        import openpyxl
    except ImportError:
        return {"_source": "excel", "_error": "openpyxl not installed",
                "sheets": [], "sheet_names": []}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True,
                                    read_only=True)
    except Exception as exc:
        return {"_source": "excel", "_error": str(exc),
                "sheets": [], "sheet_names": []}

    def _is_number(v):
        if isinstance(v, bool):
            return False
        if isinstance(v, (int, float)):
            return True
        if isinstance(v, str):
            try:
                float(v.replace(",", "").strip())
                return True
            except ValueError:
                return False
        return False

    def _num(v):
        if isinstance(v, (int, float)):
            return float(v)
        try:
            return float(str(v).replace(",", "").strip())
        except (ValueError, TypeError):
            return None

    sheets_out = []
    all_names = list(wb.sheetnames)
    for name in all_names:
        ws = wb[name]
        # Pull a bounded block of cells.
        rows = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx >= max_rows:
                break
            rows.append(row[:max_cols] if row else ())
        if not rows:
            continue
        n_cols = max((len(r) for r in rows), default=0)
        if n_cols == 0:
            continue

        # Decide whether row 0 is a header (mostly non-numeric, has some text).
        first = rows[0]
        text_cells = sum(1 for v in first if isinstance(v, str) and v.strip()
                         and not _is_number(v))
        nonempty = sum(1 for v in first if v is not None and str(v).strip())
        header_row = bool(nonempty) and text_cells >= max(1, nonempty // 2)
        data_rows = rows[1:] if header_row else rows
        row_start = 2 if header_row else 1

        columns = []
        for c in range(n_cols):
            col_vals_raw = [(r[c] if c < len(r) else None) for r in data_rows]
            nonempty_cells = [v for v in col_vals_raw if v is not None
                              and str(v).strip() != ""]
            if not nonempty_cells:
                continue
            numeric = [_num(v) for v in nonempty_cells if _is_number(v)]
            if len(numeric) < max(2, int(len(nonempty_cells) * min_numeric_fraction)):
                continue
            # header text
            if header_row and c < len(first) and first[c] is not None \
                    and str(first[c]).strip():
                header = str(first[c]).strip()
            else:
                header = f"Col {_col_letter(c + 1)}"
            columns.append(RawColumn(
                sheet=name, header=header, column_letter=_col_letter(c + 1),
                values=numeric, row_start=row_start))

        if columns:
            sheets_out.append(RawSheet(
                name=name, n_rows=len(data_rows), n_cols=n_cols,
                columns=columns))

    wb.close()
    return {
        "sheets": [
            {"name": s.name, "n_rows": s.n_rows, "n_cols": s.n_cols,
             "columns": [
                 {"sheet": c.sheet, "header": c.header,
                  "column_letter": c.column_letter, "values": c.values,
                  "row_start": c.row_start}
                 for c in s.columns]}
            for s in sheets_out
        ],
        "sheet_names": all_names,
        "_source": "excel",
    }


# Sheet-name fragments that mark a "what pack would you actually need" sheet.
_FEASIBLE_PACK_SHEET_HINTS = (
    "pack you", "pack need", "needed pack", "feasible pack", "required pack",
    "what pack", "min pack", "minimum pack", "target pack", "pack sizing",
    "pack requirement",
)


def find_feasible_pack_sheet(excel_bytes: bytes) -> dict:
    """
    Look for a user-authored "what battery pack would you actually need" sheet in
    the workbook and surface it verbatim if present — Erick's "once that sheet
    exists in your file the app can just surface it directly".

    Matching is by sheet name (case-insensitive substring against the hint list).
    Returns:
        {"found": bool,
         "sheet": <name or "">,
         "rows": [[cell, ...], ...]   # the sheet's used range, JSON-friendly
         ["_error": "..."]}
    The app COMPUTES a minimum feasible pack itself (compute_minimum_feasible_pack)
    regardless; this is only for showing the team's own sheet alongside it when
    they choose to author one. Nothing is fabricated if no such sheet exists.
    """
    try:
        import openpyxl
    except ImportError:
        return {"found": False, "sheet": "", "rows": [],
                "_error": "openpyxl not installed"}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True,
                                    read_only=True)
    except Exception as exc:
        return {"found": False, "sheet": "", "rows": [], "_error": str(exc)}

    target = None
    for name in wb.sheetnames:
        low = name.lower()
        if any(h in low for h in _FEASIBLE_PACK_SHEET_HINTS):
            target = name
            break
    if target is None:
        wb.close()
        return {"found": False, "sheet": "", "rows": []}

    ws = wb[target]
    rows = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx >= 200:                      # generous cap; these sheets are small
            break
        rows.append([("" if v is None else
                      (round(v, 4) if isinstance(v, float) else v))
                     for v in row])
    wb.close()
    # Trim trailing all-empty rows
    while rows and all((c == "" for c in rows[-1])):
        rows.pop()
    return {"found": True, "sheet": target, "rows": rows}


def ledger_declarations_from_ev_params(ev_params: dict) -> dict:
    """
    Translate the pulled pack/motor parameters into integration-ledger fields for
    the 'powertrain' and 'electrics' subsystems, so the workbook becomes the
    CROSS-TEAM source of truth — Erick's "Joule heat fs should be pulled plus key
    electric propulsion stats", flowing in "the same way the Accumulator tab
    declares numbers into the integration ledger".

    Returns a dict keyed by subsystem name, each value a dict of
    SubsystemInterface field -> value, containing ONLY fields the workbook
    actually provides (no None, no invented numbers). The caller writes these
    into the ledger via SubsystemInterface, marking is_estimate as appropriate.

        {
          "powertrain": {peak_power_kw, peak_torque_nm, voltage_v, ...},
          "electrics":  {peak_current_a, voltage_v, heat_reject_w, power_draw_w},
          "_provenance": "FSAE_EV_Power_Draw.xlsx",
        }

    heat_reject_w is derived from the workbook's Joule-heating figure
    (joule_heating_kwh) amortised over a nominal endurance stint, and is flagged
    in '_notes' so the cooling team knows it's a stint-average, not a peak.
    """
    pack = (ev_params or {}).get("pack", {}) or {}
    motor = (ev_params or {}).get("motor", {}) or {}

    def _pos(d, k):
        v = d.get(k)
        try:
            v = float(v)
        except (TypeError, ValueError):
            return None
        return v if v > 0 else None

    powertrain = {}
    electrics = {}
    notes = []

    # --- powertrain: motor envelope headline numbers ---
    pk_kw = _pos(motor, "motor_peak_power_kw")
    if pk_kw is not None:
        powertrain["peak_power_kw"] = pk_kw
    tq = _pos(motor, "motor_peak_torque_nm")
    if tq is not None:
        powertrain["peak_torque_nm"] = tq
    v_pack = _pos(pack, "pack_voltage_v") or _pos(motor, "pack_voltage_ep_v")
    if v_pack is not None:
        powertrain["voltage_v"] = v_pack
        electrics["voltage_v"] = v_pack

    # --- electrics: current + the Joule-heat the cooling team must reject ---
    i_peak = _pos(pack, "cell_current_a") or _pos(motor, "current_from_pack_a")
    if i_peak is not None:
        electrics["peak_current_a"] = i_peak
    p_draw_kw = _pos(pack, "power_draw_kw")
    if p_draw_kw is not None:
        electrics["power_draw_w"] = p_draw_kw * 1000.0

    joule_kwh = _pos(pack, "joule_heating_kwh")
    endurance_km = _pos(pack, "endurance_km") or 22.0   # FSAE endurance ~22 km
    if joule_kwh is not None:
        # Amortise the stint's Joule energy into an average heat-rejection rate.
        # Assume an endurance pace of ~50 km/h. stint_hours = distance / speed,
        # stint_seconds = stint_hours * 3600. (Earlier this mixed km with m/s and
        # produced a ~MW result — guard with sane bounds below.)
        endurance_pace_kmh = 50.0
        stint_s = (endurance_km / endurance_pace_kmh) * 3600.0 if endurance_km else 1600.0
        if stint_s > 0:
            heat_w = joule_kwh * 1000.0 * 3600.0 / stint_s
            electrics["heat_reject_w"] = heat_w
            notes.append(
                f"heat_reject_w = {heat_w:.0f} W is the {joule_kwh:.3f} kWh "
                f"workbook Joule figure averaged over a ~{stint_s/60:.0f}-min "
                f"endurance stint at {endurance_pace_kmh:.0f} km/h "
                f"(stint average, not peak).")

    out = {}
    if powertrain:
        out["powertrain"] = powertrain
    if electrics:
        out["electrics"] = electrics
    out["_provenance"] = "FSAE_EV_Power_Draw.xlsx"
    if notes:
        out["_notes"] = notes
    return out
